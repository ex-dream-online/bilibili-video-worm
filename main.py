# coding=utf-8
import requests
import re
import datetime
import random
import time
import sys
import openpyxl
import os
import numpy


global bv,header

header={
    'Host':'api.bilibili.com',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    'Cookie':''
}

# 验证bv号是否正确
def get_code_judge():
    urla='http://api.bilibili.com/x/web-interface/archive/stat?bvid='
    js=requests.session().get(urla+bv,headers=header).text
    raw=re.search(r'"code":.*?,',str(js)).group(0)
    code=re.search(r'\d+',str(raw)).group(0)
    if int(code)!=0:
        print('您输入的不是BV号,程序结束')
        sys.exit()
# 获取av号
def get_aid():
    global aid
    urla='http://api.bilibili.com/x/web-interface/archive/stat?bvid='
    js=requests.session().get(urla+bv,headers=header).text
    # 数据正则筛选
    raw=re.search(r'"aid":\d+',str(js)).group(0)
    aid=re.search(r'\d+',str(raw)).group(0)
    return aid
# 获取cid
def get_cid():
    global cid
    urla='https://api.bilibili.com/x/player/pagelist?bvid='
    urlb='&jsonp=jsonp'
    js=requests.session().get(urla+bv+urlb,headers=header).text
    # 数据正则筛选
    raw=re.search(r'"cid":\d+',str(js)).group(0)
    cid=re.search(r'\d+',str(raw)).group(0)
    return cid
# 获取视频标题
def get_title():
    global title
    urla='https://api.bilibili.com/x/web-interface/view?bvid='
    js=requests.session().get(urla+bv,headers=header).text
    js=eval("u"+"\'"+js+"\'")# 处理编码问题
    # 数据正则
    raw=re.search(r'"title":".*?"',str(js)).group(0)
    raw1=re.split(r'"',raw)
    title=raw1[3]
    return title
# 获取视频数据 播放量，弹幕数，回复数，点赞数，硬币数，分享数，点赞数
def get_data():
    global view,danmaku,reply,favorite,coin,share,like
    urla='http://api.bilibili.com/x/web-interface/archive/stat?bvid='
    js=requests.session().get(urla+bv,headers=header).text
    # 数据筛选
    # 播放量
    raw_view=re.search(r'"view":\d+',str(js)).group(0)
    view=re.search(r'\d+',str(raw_view)).group(0)
    # 弹幕数
    raw_danmaku=re.search(r'"danmaku":\d+',str(js)).group(0)
    danmaku=re.search(r'\d+',str(raw_danmaku)).group(0)
    # 回复数
    raw_reply=re.search(r'"reply":\d+',str(js)).group(0)
    reply=re.search(r'\d+',str(raw_reply)).group(0)
    # 收藏数
    raw_favorite=re.search(r'"favorite":\d+',str(js)).group(0)
    favorite=re.search(r'\d+',str(raw_favorite)).group(0)
    # 硬币数
    raw_coin=re.search(r'"coin":\d+',str(js)).group(0)
    coin=re.search(r'\d+',str(raw_coin)).group(0)
    # 分享数
    raw_share=re.search(r'"share":\d+',str(js)).group(0)
    share=re.search(r'\d+',str(raw_share)).group(0)
    # 点赞数
    raw_like=re.search(r'"like":\d+',str(js)).group(0)
    like=re.search(r'\d+',str(raw_like)).group(0)
    return view
# 获取当前观看人数
def get_concurrent_viewers():
    global concurrent_viewers
    urla='https://api.bilibili.com/x/player/online/total?aid='
    urlb='&cid='
    urlc='&bvid='
    js=requests.session().get(urla+aid+urlb+cid+urlc+bv,headers=header).text
    # 数据正则筛选
    raw=re.search(r'"total":"\d+',str(js)).group(0)
    concurrent_viewers=re.search(r'\d+',str(raw)).group(0)
    if int(concurrent_viewers)>=1000:#判断当前观看人数是否大于1000，需不需要加+号
        concurrent_viewers=concurrent_viewers+'+'
        return concurrent_viewers
    return concurrent_viewers
# 获取查询时间
def get_concurrent_time():
    global time1
    tm=datetime.datetime.now()
    time1=tm.strftime("%Y-%m-%d %H:%M:%S")
    return time1
# 建立新Excel
def create_excel_xlsx(path,sheet_name='sheet1'):
    workbook=openpyxl.Workbook()
    sheet=workbook.active
    sheet.title=sheet_name
    sheet.append(['播放量', '评论数', '分享数', '点赞量', '当前观看人数', '数据记录时间'])
    workbook.save(path)
# 写入Excel
def write_excel_xlsx_append(path,value,truncate_sheet=False):
    if not os.path.exists(path):
        create_excel_xlsx(path)

    workbook=openpyxl.load_workbook(path)
    # list转换为array
    value=numpy.array(value)
    # 取第一张表
    sheetnames=workbook.sheetnames
    sheet=workbook[sheetnames[0]]
    sheet=workbook.active
    # 获得行数
    startrows=sheet.max_row
    index=len(value)
    # 按数据坐标将数据写入excel
    # index+1是让列从1开始，因为python里从0开始，Excel里从1开始
    for i in range(1,index+1):
        sheet.cell(row=startrows+1,column=i).value=value[i-1]
    workbook.save(path)


bv=input('请输入您想要爬取的视频的BV号:')
get_code_judge()
n=int(input('请输入您想爬取的数据条数(每两条数据间隔大概30s,您可以在下方更改间隔时间):'))
get_aid();get_cid();get_title()
for i in range(n):
    #延时
    if i>0:
        time.sleep(40+random.randint(-7,15))
    try:
        get_data()
        get_concurrent_viewers()
    except Exception:
        print('Zzzzzz...')
        time.sleep(600)
        get_data()
        get_concurrent_viewers()
    get_concurrent_time()
    print('播放量:%d'%int(view),'评论数:%d'%int(reply),'分享数:%d'%int(share),'点赞数:%d'%int(like),
          '当前观看人数:%s'%concurrent_viewers,'数据记录时间:%s'%time1,'剩余抓取次数:%d'%int(n-i-1))
    xx_info=[view,reply,share,like,concurrent_viewers,time1]
    write_excel_xlsx_append(title+'爬取数据'+'.xlsx',xx_info)
print('数据爬取完毕')
